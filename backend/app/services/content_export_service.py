"""Export content queue (timeline drafts) to Excel for review before approval."""

from __future__ import annotations

import io
from datetime import UTC, date, datetime
from pathlib import Path
from uuid import UUID

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from PIL import Image as PILImage
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import Settings, get_settings


def _photo_download_url(photo_id: str | None, client_id: UUID, settings: Settings) -> str:
    if not photo_id:
        return ""
    path = f"/api/v1/gbp/photos/{photo_id}/file"
    base = (settings.public_api_base_url or "").strip().rstrip("/")
    return f"{base}{path}" if base else path


async def build_content_queue_xlsx(session: AsyncSession, client_id: UUID) -> bytes:
    settings = get_settings()
    rows = (
        await session.execute(
            text(
                """
                SELECT id, content_type, status, payload, generated_at
                FROM rp_content_queue
                WHERE client_id = :cid
                  AND status IN ('pending', 'approved')
                  AND content_type IN ('gbp_post', 'landing_page')
                ORDER BY
                  COALESCE((payload->>'week_number')::int, 99),
                  COALESCE(payload->>'scheduled_for', ''),
                  generated_at
                """
            ),
            {"cid": str(client_id)},
        )
    ).mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Content Plan"

    headers = [
        "Week",
        "Scheduled",
        "Type",
        "Status",
        "Title",
        "Keyword",
        "Suburb",
        "Volume",
        "KD",
        "Word count",
        "Body",
        "Photo ID",
        "Image download URL",
        "Notes",
        "Item ID",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in rows:
        payload = r["payload"] if isinstance(r["payload"], dict) else {}
        if not isinstance(payload, dict):
            payload = {}
        photo_id = str(payload.get("photo_id") or "").strip() or None
        ws.append(
            [
                payload.get("week_number"),
                payload.get("scheduled_for"),
                r["content_type"],
                r["status"],
                payload.get("title"),
                payload.get("target_keyword"),
                payload.get("suburb"),
                payload.get("keyword_volume"),
                payload.get("keyword_difficulty"),
                payload.get("word_count"),
                payload.get("body"),
                photo_id or "",
                _photo_download_url(photo_id, client_id, settings),
                payload.get("notes"),
                str(r["id"]),
            ]
        )

    # Auto-width (approximate)
    for col in ws.columns:
        max_len = 12
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, min(80, len(str(cell.value))))
        ws.column_dimensions[col_letter].width = max_len

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def content_export_filename() -> str:
    stamp = datetime.now(UTC).strftime("%Y-%m-%d")
    return f"rankpilot-content-plan-{stamp}.xlsx"


def _find_photo_file(photo_id: str | None, client_id: UUID) -> Path | None:
    """Locate the stored post image on disk by photo id."""
    if not photo_id:
        return None
    base = Path("uploads/gbp") / str(client_id)
    for ext in (".png", ".jpg", ".jpeg", ".webp", ".gif"):
        p = base / f"{photo_id}{ext}"
        if p.is_file():
            return p
    return None


def _build_thumbnail(path: Path, max_w: int = 130, max_h: int = 100) -> tuple[io.BytesIO, int, int] | None:
    """Return (PNG buffer, width_px, height_px) for an embeddable thumbnail."""
    try:
        with PILImage.open(path) as im:
            im = im.convert("RGB")
            w, h = im.size
            scale = min(max_w / w, max_h / h, 1.0)
            new_w = max(1, int(w * scale))
            new_h = max(1, int(h * scale))
            im = im.resize((new_w, new_h))
            out = io.BytesIO()
            im.save(out, format="PNG")
            out.seek(0)
            return out, new_w, new_h
    except Exception:
        return None


async def build_gbp_posts_xlsx(
    session: AsyncSession,
    client_id: UUID,
    *,
    post_ids: list[str] | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
) -> bytes:
    """Excel of GBP posts (content, keyword, status, schedule + embedded image).

    Optional filters:
    - post_ids: export only these queue row IDs
    - date_from / date_to: filter on generated_at (fallback created_at)
    """
    settings = get_settings()
    conditions = [
        "client_id = :cid",
        "content_type = 'gbp_post'",
        "status <> 'archived'",
    ]
    params: dict = {"cid": str(client_id)}

    if post_ids:
        conditions.append("id::text = ANY(CAST(:post_ids AS text[]))")
        params["post_ids"] = post_ids

    if date_from:
        conditions.append("DATE(COALESCE(generated_at, created_at)) >= :date_from")
        params["date_from"] = date_from

    if date_to:
        conditions.append("DATE(COALESCE(generated_at, created_at)) <= :date_to")
        params["date_to"] = date_to

    where_sql = " AND ".join(conditions)
    rows = (
        await session.execute(
            text(
                f"""
                SELECT id, status, payload, generated_at, published_at, created_at
                FROM rp_content_queue
                WHERE {where_sql}
                ORDER BY
                  COALESCE(payload->>'scheduled_for', '') ASC,
                  created_at ASC
                """
            ),
            params,
        )
    ).mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "GBP Posts"

    headers = [
        "#",
        "Image",
        "Title",
        "Target keyword",
        "Status",
        "Scheduled date",
        "Word count",
        "Post content",
        "Image download URL",
        "Generated at",
        "Published at",
        "Post ID",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    image_col_letter = "B"  # "Image" column
    for idx, r in enumerate(rows, start=1):
        payload = r["payload"] if isinstance(r["payload"], dict) else {}
        if not isinstance(payload, dict):
            payload = {}
        photo_id = str(payload.get("photo_id") or "").strip() or None
        gen = r["generated_at"].strftime("%Y-%m-%d %H:%M") if r["generated_at"] else ""
        pub = r["published_at"].strftime("%Y-%m-%d %H:%M") if r["published_at"] else ""
        ws.append(
            [
                idx,
                "",  # image is added as a floating picture below
                payload.get("title"),
                payload.get("target_keyword"),
                str(r["status"]).replace("_", " "),
                payload.get("scheduled_for"),
                payload.get("word_count"),
                payload.get("body"),
                _photo_download_url(photo_id, client_id, settings),
                gen,
                pub,
                str(r["id"]),
            ]
        )

        excel_row = idx + 1  # +1 for header row
        photo_path = _find_photo_file(photo_id, client_id)
        thumb = _build_thumbnail(photo_path) if photo_path else None
        if thumb:
            buf_img, w_px, h_px = thumb
            xl_img = XLImage(buf_img)
            xl_img.width = w_px
            xl_img.height = h_px
            xl_img.anchor = f"{image_col_letter}{excel_row}"
            ws.add_image(xl_img)
            # Size the row so the thumbnail fits (≈0.75pt per px).
            ws.row_dimensions[excel_row].height = max(20, h_px * 0.75)

    # Column widths (the Image column is set wide enough for thumbnails).
    width_by_header = {
        "#": 5,
        "Image": 20,
        "Title": 32,
        "Target keyword": 22,
        "Status": 14,
        "Scheduled date": 14,
        "Word count": 11,
        "Post content": 80,
        "Image download URL": 30,
        "Generated at": 17,
        "Published at": 17,
        "Post ID": 38,
    }
    for col_idx, header in enumerate(headers, start=1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = width_by_header.get(header, 16)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def gbp_posts_export_filename(
    *,
    date_from: date | None = None,
    date_to: date | None = None,
    count: int | None = None,
) -> str:
    stamp = datetime.now(UTC).strftime("%Y-%m-%d")
    parts = ["rankpilot-gbp-posts", stamp]
    if date_from and date_to:
        parts.append(f"{date_from.isoformat()}-to-{date_to.isoformat()}")
    elif date_from:
        parts.append(f"from-{date_from.isoformat()}")
    elif date_to:
        parts.append(f"until-{date_to.isoformat()}")
    if count is not None:
        parts.append(f"{count}-posts")
    return "-".join(parts) + ".xlsx"
